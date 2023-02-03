import openpyxl
from openpyxl import load_workbook

# Load the existing Excel file
wb = openpyxl.load_workbook('report.xlsx')

# Iterate through each sheet in the Excel file
for sheet in wb:
    # Iterate through each column in the sheet
    for i in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=i).value is not None:
            if "Last" not in sheet.cell(row=1, column=i).value:
                # Get the current column header value
                current_header_value = sheet.cell(row=1, column=i).value
                        # Update the column header value by adding the string
                sheet.cell(row=1, column=i).value = "Last_" + current_header_value
            # Iterate through all the rows and columns of the sheet

# Save the changes to the existing Excel file

    wb.save('report.xlsx')


