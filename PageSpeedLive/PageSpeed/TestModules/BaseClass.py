import time

import openpyxl as xl
import requests
from openpyxl.styles import PatternFill

from PageSpeed.TestModules.dbConnection import dbConnection

class commonFunction:
    global row
    global col
    global flagF
    row = 2
    col = 1
    flagF = 0
    overall_score_mob = 0
    accessibility_score_mob = 0
    seo_score_mob = 0
    bestprts_score_mob = 0
    overall_score_desk = 0
    accessibility_score_desk = 0
    seo_score_desk = 0
    bestprts_score_desk = 0

    def data_result(self, count, sheet_Name, line, performance, accessibilty, seo, bstpract, performance_des,
                    accessibilty_des, seo_des,
                    bstpract_des):

        global row
        global col
        global flagF

        wb = xl.load_workbook('PageSpeedReport1.xlsx')
        print(row, col)
        print("Count is -****", count)

        if count == 2:
            if sheet_Name == 'Home':
                row = 2
                col = 1
            else:
                print("Nothin update in row and col")
        if count == 3:
            if sheet_Name == 'Timeline':
                row = 2
                col = 1
            else:
                print("Nothin update in row and col")
        if count == 4:
            if sheet_Name == 'Gallery':
                row = 2
                col = 1
            else:
                print("Nothin update in row and col")
        if count == 5:
            if sheet_Name == 'Map':
                row = 2
                col = 1
            else:
                print("Nothin update in row and col")
        if count == 6:
            if sheet_Name == 'Contact':
                row = 2
                col = 1
            else:
                print("Nothin update in row and col")
        if count == 7:
            if sheet_Name == 'Additional':
                row = 2
                col = 1
            else:
                print("Nothing update in row and col")

        time.sleep(1)
        sheet_N = wb[sheet_Name]
        wb.active = wb.index(sheet_N)

        redFill = PatternFill(start_color='FFFF0000',
                              end_color='FFFF0000',
                              fill_type='solid')

        sheet_N.cell(row, col, 'Mobile')
        sheet_N.cell(row, col + 1, line)
        if performance < 80:
            sheet_N.cell(row, col + 2, performance).fill = redFill
        else:
            sheet_N.cell(row, col + 2, performance)
        if accessibilty < 95:
            sheet_N.cell(row, col + 3, accessibilty).fill = redFill
        else:
            sheet_N.cell(row, col + 3, accessibilty)
        if seo < 95:
            sheet_N.cell(row, col + 4, seo).fill = redFill
        else:
            sheet_N.cell(row, col + 4, seo)
        if bstpract < 80:
            sheet_N.cell(row, col + 5, bstpract).fill = redFill
        else:
            sheet_N.cell(row, col + 5, bstpract)
        row = row + 1
        print(row)
        sheet_N.cell(row, col, 'desktop')
        sheet_N.cell(row, col + 1, line)
        if performance_des < 80:
            sheet_N.cell(row, col + 2, performance_des).fill = redFill
        else:
            sheet_N.cell(row, col + 2, performance_des)
        if accessibilty_des < 95:
            sheet_N.cell(row, col + 3, accessibilty_des).fill = redFill
        else:
            sheet_N.cell(row, col + 3, accessibilty_des)
        if seo_des < 95:
            sheet_N.cell(row, col + 4, seo_des).fill = redFill
        else:
            sheet_N.cell(row, col + 4, seo_des)
        if bstpract_des < 80:
            sheet_N.cell(row, col + 5, bstpract_des).fill = redFill
        else:
            sheet_N.cell(row, col + 5, bstpract_des)
        row = row + 1
        print(row)

        wb.save('PageSpeedReport1.xlsx')

    def Mobile_apidata(self, count, sheet_Name, line):
        global overall_score_mob
        global accessibility_score_mob
        global seo_score_mob
        global bestprts_score_mob
        global overall_score_desk
        global accessibility_score_desk
        global seo_score_desk
        global bestprts_score_desk
        try:
            x = f'https://www.googleapis.com/pagespeedonline/v5/runPagespeed?key=AIzaSyA0BaRRoglLAmGgPgRaGJWGfyuCqKCBwOk&url={line}&category=ACCESSIBILITY&category=best-practices&category=PERFORMANCE&category=PWA&category=SEO&strategy=mobile'
            print(f'Requesting {x}...')
            Mob_data = requests.get(x).json()
            y = f'https://www.googleapis.com/pagespeedonline/v5/runPagespeed?key=AIzaSyA0BaRRoglLAmGgPgRaGJWGfyuCqKCBwOk&url={line}&category=ACCESSIBILITY&category=best-practices&category=PERFORMANCE&category=PWA&category=SEO&strategy=desktop'
            print(f'Requesting {y}...')
            response_desk = requests.get(y)
            Desk_data = response_desk.json()
            print(requests.get(x).status_code)
            if requests.get(x).status_code == 200 and requests.get(y).status_code == 200:
                print("If enter")
                overall_score_mob = Mob_data["lighthouseResult"]["categories"]["performance"]["score"] * 100
                accessibility_score_mob = Mob_data["lighthouseResult"]["categories"]["accessibility"]["score"] * 100
                seo_score_mob = Mob_data["lighthouseResult"]["categories"]["seo"]["score"] * 100
                bestprts_score_mob = Mob_data["lighthouseResult"]["categories"]["best-practices"]["score"] * 100
                print("Mobile Overall Performance -", overall_score_mob, "\n" "Mobile Accessibility Score - ",
                      accessibility_score_mob, "\n" "Mobile SEO Score - ", seo_score_mob,
                      "\n" "Mobile Best-practices Score - ",
                      bestprts_score_mob)
                time.sleep(1)
                overall_score_desk = Desk_data["lighthouseResult"]["categories"]["performance"]["score"] * 100
                accessibility_score_desk = Desk_data["lighthouseResult"]["categories"]["accessibility"]["score"] * 100
                seo_score_desk = Desk_data["lighthouseResult"]["categories"]["seo"]["score"] * 100
                bestprts_score_desk = Desk_data["lighthouseResult"]["categories"]["best-practices"]["score"] * 100
                print("Desktop Overall Performance -", overall_score_desk, "\n" "Desktop Accessibility Score - ",
                      accessibility_score_desk, "\n" "Desktop SEO Score - ", seo_score_desk,
                      "\n" "desktop Best-practices Score - ", bestprts_score_desk)
                time.sleep(1)
                if sheet_Name == sheet_Name:
                    self.data_result(count, sheet_Name, line, overall_score_mob, accessibility_score_mob, seo_score_mob,
                                     bestprts_score_mob, overall_score_desk, accessibility_score_desk,
                                     seo_score_desk, bestprts_score_desk)

                    dbConnection.dbconn("Mobile",sheet_Name,line,overall_score_mob, accessibility_score_mob, seo_score_mob,
                                     bestprts_score_mob)
                    dbConnection.dbconn("Website", sheet_Name, line, overall_score_desk, accessibility_score_desk,
                                     seo_score_desk, bestprts_score_desk)


            else:
                print("else enter")
                self.data_result(count, sheet_Name, line, 0, 0, 0, 0, 0, 0, 0, 0)
                dbConnection.dbconn("Mobile", sheet_Name, line, overall_score_mob, accessibility_score_mob,
                                    seo_score_mob,
                                    bestprts_score_mob)
                dbConnection.dbconn("Website", sheet_Name, line, overall_score_desk, accessibility_score_desk,
                                    seo_score_desk, bestprts_score_desk)
                #dbConnection()
        except:
            print("except enter")
            self.data_result(count, sheet_Name, line, 0, 0, 0, 0, 0, 0, 0, 0)
            dbConnection.dbconn("Mobile", sheet_Name, line,0,0,0,0)
            dbConnection.dbconn("Website", sheet_Name, line, 0,0,0,0)
            pass
