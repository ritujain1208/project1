from datetime import time, datetime
from os import path

import openpyxl
import pytest
from openpyexcel import Workbook
from self import self


# @pytest.fixture(autouse=True, scope="class")
def sheetCreate():
    wb = Workbook()
    print("Sheet creating")
    ws1 = wb.create_sheet("Sheet_A", 0)
    ws1.title = "Store"
    ws2 = wb.create_sheet("Sheet_B", 1)
    ws2.title = "Home"
    ws3 = wb.create_sheet("Sheet_C", 2)
    ws3.title = "Timeline"
    ws4 = wb.create_sheet("Sheet_D", 3)
    ws4.title = "Gallery"
    ws5 = wb.create_sheet("Sheet_E", 4)
    ws5.title = "Map"
    ws6 = wb.create_sheet("Sheet_F", 5)
    ws6.title = "Contact"
    ws7 = wb.create_sheet("Sheet_G", 6)
    ws7.title = "Additional"
    default_sheet = wb['Sheet']
    wb.remove(default_sheet)
    for sheetName in wb.sheetnames:
        col_name = ['type', 'urls', 'Performance', 'Accessibility',
                    'SEO', 'Best Practice']
        wb[sheetName].append(col_name)

    wb.save('PageSpeedReport' + "1" + '.xlsx')
