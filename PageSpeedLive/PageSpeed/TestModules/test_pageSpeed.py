import shutil
import time

import openpyxl
import pandas as pd
from openpyexcel import load_workbook
from openpyexcel.styles import PatternFill

from PageSpeed.TestModules.BaseClass import commonFunction
from PageSpeed.TestModules.conftest import sheetCreate
from PageSpeed.TestModules.mailreport import mail_send
from copy import copy


class TestPagespeed(commonFunction):

    def test_pageSpeedApi(self):
        sheetCreate()
        print("Enter")
        df = pd.ExcelFile(r"readurls.xlsx")
        count = 0
        sum = 0
        print("next")
        for sheet_Name in df.sheet_names:
            df = pd.read_excel(r"readurls.xlsx", sheet_name=sheet_Name, index_col=None)
            url = df['urls'].tolist()
            num = len(url)
            print("no of urls - " + sheet_Name + "--", num)
            print("count is ", count)
            sum = sum + 1

            for line in url:
                time.sleep(1)
                print("value of count is - ", count)
                commonFunction.Mobile_apidata(self, count, sheet_Name, line)
                print("all data get")
                count = 0
            count = sum + 1
            print(sum, "-", count)

    def test_compare(self):
        print("test compare")
        wb1 = load_workbook('PageSpeedReport1.xlsx')
        wb2 = load_workbook('report.xlsx')

        for worksheet in wb1.sheetnames:
            sheet1 = wb1[worksheet]
            sheet2 = wb2[worksheet]

            # iterate through the rows and columns of both worksheets
            for row in range(2, sheet1.max_row + 1):
                for col in range(3, sheet1.max_column + 1):
                    cell1 = sheet1.cell(row, col)
                    cell2 = sheet2.cell(row, col)
                    if cell1.value is not None and cell2.value is not None:
                        if cell1.value > cell2.value:
                            print("enter in c1>c2")
                            fill_cell1 = PatternFill(start_color='00FF00',
                                                     end_color='00FF00',
                                                     fill_type='solid')
                            ws = wb1["{0}".format(worksheet)]
                            ws.cell(row, col).fill = fill_cell1
                        if cell1.value < cell2.value:
                            print("enter in c1<c2")
                        if cell1.value == cell2.value:
                            print("enter in c1==c2")
                    else:
                        print("Value is blank")
        wb1.save(filename="PageSpeedReport1.xlsx")

    #
    # # run the script again
    def test_moveCodeToMergeFile(self):
        existing_workbook = load_workbook("combineReport.xlsx")
        other_workbook = load_workbook("PageSpeedReport1.xlsx")
        for sheet in other_workbook:
            # Check if the sheet exists in the existing workbook
            if sheet.title in existing_workbook.sheetnames:
                # Get the existing worksheet
                existing_worksheet = existing_workbook[sheet.title]
                # Get the last column letter in the existing worksheet
                last_column = existing_worksheet.max_column
                # Iterate over each row and column of the other sheet
                for i in range(1, sheet.max_row + 1):
                    for j in range(1, sheet.max_column + 1):
                        # Get the cell value
                        value = sheet.cell(row=i, column=j).value
                        # Get the cell style
                        style = sheet.cell(row=i, column=j)._style
                        fill = copy(sheet.cell(row=i, column=j).fill)
                        # Write the value to the next column of the existing worksheet
                        existing_worksheet.cell(row=i, column=j).value = value
                        # Apply the cell style to the existing worksheet
                        existing_worksheet.cell(row=i, column=j)._style = style
                        existing_worksheet.cell(row=i, column=j).fill = fill
                        # Apply the
        existing_workbook.save("combineReport.xlsx")

    def test_moveCodeLastMonth(self):
        time.sleep(3)
        existing_workbook = load_workbook("combineReport.xlsx")
        other_workbook = load_workbook("report.xlsx")

        for sheet in other_workbook:
            # Check if the sheet exists in the existing workbook
            if sheet.title in existing_workbook.sheetnames:
                # Get the existing worksheet
                existing_worksheet = existing_workbook[sheet.title]
                # Get the last column letter in the existing worksheet
                last_column = existing_worksheet.max_column
                # Iterate over each row and column of the other sheet
                for i in range(1, sheet.max_row + 1):
                    for j in range(2, sheet.max_column + 1):
                        # Get the cell value
                        value = sheet.cell(row=i, column=j).value
                        # Get the cell style
                        style = sheet.cell(row=i, column=j)._style
                        fill = copy(sheet.cell(row=i, column=j).fill)

                        # Write the value to the next column of the existing worksheet
                        existing_worksheet.cell(row=i, column=last_column + j).value = value
                        # Apply the cell style to the existing worksheet
                        existing_worksheet.cell(row=i, column=last_column + j)._style = style
                        existing_worksheet.cell(row=i, column=last_column + j).fill = fill

                        # Apply the
        existing_workbook.save("combineReport.xlsx")

    def test_mail(self):
        print("mail code")
        mail_send("combineReport.xlsx")
        time.sleep(2)

    def test_deleteFiles(self):
        # move data of file one to aother
        source = "PageSpeedReport1.xlsx"
        destination = "report.xlsx"
        #
        shutil.copy(source, destination)
        # #     # if os.path.isfile(source):
        # #     #     os.remove(source)
        # #     #     print("File removed")
        # #
        # #  # Load the existing Excel file
        wb = load_workbook(destination)
        # Iterate through each sheet in the Excel file
        for sheet in wb:
            # Iterate through each column in the sheet
            for i in range(1, sheet.max_column + 1):
                if sheet.cell(row=1, column=i).value is not None:
                    if "Last" not in sheet.cell(row=1, column=i).value:
                        # Get the current column header value
                        current_header_value = sheet.cell(row=1, column=i).value
                        # Update the column header value by adding the string
                        sheet.cell(row=1, column=i).value = "Last_" + current_header_value
                # Iterate through all the rows and columns of the sheet
        wb.save(destination)
        # # Open the existing workbook

        workbook = load_workbook(destination)
        fill_cell1 = PatternFill(start_color='00FF00',
                                                         end_color='00FF00',
                                                         fill_type='solid')
        # Iterate through all the sheets in the workbook
        for sheet in workbook:
            # Iterate through all the rows and columns of the sheet
            for i in range(2, sheet.max_row + 1):
                for j in range(2, sheet.max_column + 1):
                    if sheet.cell(row=i, column=j).value is not None:
                        if sheet.cell(row=i, column=j).fill == fill_cell1:
                            print("^^^^^^^^^^^^")
                            sheet.cell(row=i,column=j).fill = PatternFill(
                                                     fill_type=None)
                            # Get the current column header value
        # Save the workbook
        workbook.save(destination)


        # Save the changes to the existing Excel file

    # def test_afterMail(self):
    #     workbook = openpyxl.load_workbook('combineReport.xlsx')
    #
    #     # Iterate through all the sheets in the workbook
    #     for sheet in workbook:
    #         # Delete all the rows in the sheet
    #         for row in sheet.iter_rows():
    #             for cell in row:
    #                 cell.value = None
    #                 cell.fill = openpyxl.styles.PatternFill(fill_type=None)
    #         # Save the workbook
    #         workbook.save('combineReport.xlsx')
    #     # pagespeed report

    # # def test_schedule(self, i=1):
    # #     print("task scheduling")
    # #     # copy excel report to another report which genrated:------------------
    # #     time.sleep(1)
    # #     # # calculate current time of system:---------------
    # #     time_now = datetime.now()
    # #     current_time = time_now.strftime("%H:%M:%S")
    # #     print("The current date and time is", current_time)
    # #     updated_time = datetime.now() + timedelta(minutes=30)
    # #     print("Updated time", updated_time)
    # #     print("Done")
    # #     print('Wait time over')
    # #     # else: 3600
    # #     time.sleep(1800)
    # #     print(datetime.now())
    ##  self.test_pageSpeedApi()#
    #
